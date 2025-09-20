import io
import os
import re
import sys
import glob
from datetime import datetime

from openpyxl import load_workbook, Workbook

# ===================== CONFIG =====================
BASE_DIR   = os.path.dirname(__file__)
EXCEL_PATH = os.path.join(BASE_DIR, "data.xlsx")
SHEET_NAME = "Sheet1"
CREDS_PATH = os.path.join(BASE_DIR, "DLOCR.json")

DEBUG    = True   # print OCR text to Terminal
SAVE_RAW = True   # also write last_ocr.txt
# ==================================================

# Quiet gRPC warnings a bit
os.environ.setdefault("GRPC_VERBOSITY", "ERROR")
os.environ.setdefault("GRPC_TRACE", "")

# ---------------- Google Vision -------------------
from google.cloud import vision
from google.oauth2 import service_account

if os.path.exists(CREDS_PATH):
    _creds = service_account.Credentials.from_service_account_file(CREDS_PATH)
    client = vision.ImageAnnotatorClient(credentials=_creds)
else:
    client = vision.ImageAnnotatorClient()
# --------------------------------------------------


# ================= Excel helpers ==================
HEADERS = [
    "Phone", "Date", "Have you Been Here Before?", "First Name", "Last Name",
    "Date Of Birth", "Email Address", "Address", "City", "State", "Zip Code",
    "Type Of Service", "Payment Type", "Amount", "Comments", "View Image"
]

def _sheet_has_real_data(ws):
    for row in ws.iter_rows(values_only=True):
        if any(v not in (None, "") for v in row):
            return True
    return False

def ensure_workbook():
    """Create workbook if missing, or re-header a blank/ghosted sheet."""
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)
        return

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    ws.title = SHEET_NAME

    if not _sheet_has_real_data(ws):
        ws.delete_rows(1, ws.max_row or 1)
        ws.append(HEADERS)
    else:
        existing = [c.value for c in ws[1]]
        if existing != HEADERS:
            ws.delete_rows(1, 1)
            ws.insert_rows(1)
            for i, h in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=i, value=h)
    wb.save(EXCEL_PATH)

def _first_truly_empty_row(ws, min_cols):
    """Return first data row (>=2) where all checked cells are empty."""
    r = 2
    ncols = max(ws.max_column, min_cols)
    while True:
        if all((ws.cell(row=r, column=c).value in (None, "")) for c in range(1, ncols + 1)):
            return r
        r += 1

def append_row(values: dict, image_path: str):
    """Append values into first truly empty row; add hyperlink in 'View Image'."""
    ensure_workbook()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    row = _first_truly_empty_row(ws, len(HEADERS))
    for col, header in enumerate(HEADERS, start=1):
        if header == "View Image":
            cell = ws.cell(row=row, column=col, value="View Image")
            cell.hyperlink = image_path
            cell.style = "Hyperlink"
        else:
            ws.cell(row=row, column=col, value=values.get(header, ""))
    wb.save(EXCEL_PATH)
# ==================================================


# ================= OCR & parsing ==================
def find_latest_image(directory):
    files = glob.glob(os.path.join(directory, "*.[jp][pn]g"))  # .jpg/.jpeg/.png
    if not files:
        raise FileNotFoundError("No image files found in the directory.")
    return max(files, key=os.path.getctime)

def detect_text(image_path) -> str:
    """Return the full OCR text (Vision block #0) as a single string."""
    with io.open(image_path, "rb") as f:
        content = f.read()
    image = vision.Image(content=content)
    response = client.text_detection(image=image)
    if response.error.message:
        raise RuntimeError(response.error.message)
    full = response.text_annotations[0].description if response.text_annotations else ""

    if DEBUG:
        print("\n===== OCR RAW TEXT START =====\n", flush=True)
        print(full, flush=True)
        print("\n===== OCR RAW TEXT END =====\n", flush=True)
    if SAVE_RAW:
        with open(os.path.join(BASE_DIR, "last_ocr.txt"), "w", encoding="utf-8") as f:
            f.write(full)
    return full

def _extract_names(full_text: str):
    """
    Extract First/Last from common DL patterns, avoiding city/state lines.
    Order:
      1) Numeric fields: '1 LAST', '2 FIRST'
      2) 'LAST, FIRST [MIDDLE]'
      3) Labeled LN/FN
      4) Uppercase line near the top (LAST FIRST [MIDDLE]), skipping commas/digits/state codes.
    """
    first = last = ""

    # 1) Numeric fields (very common)
    m1 = re.search(r"\b1\s+([A-Z][A-Z'`-]+)\b", full_text)  # last
    m2 = re.search(r"\b2\s+([A-Z][A-Z'`-]+)\b", full_text)  # first
    if m1: last = m1.group(1).title()
    if m2: first = m2.group(1).title()
    if first and last:
        return first, last

    # 2) 'LAST, FIRST [MIDDLE]'
    m = re.search(r"\b([A-Z][A-Z'`-]+)\s*,\s*([A-Z][A-Z'`-]+)(?:\s+[A-Z][A-Z'`-]+)?\b", full_text)
    if m:
        return (m.group(2).title(), m.group(1).title())

    # 3) Labeled LN/FN
    m = re.search(r"(Last Name|Surname|Family Name|LN)[:\s]+([A-Za-z'`-]+)", full_text, re.I)
    if m and not last:
        last = m.group(2).title()
    m = re.search(r"(First Name|Given Names|FN)[:\s]+([A-Za-z'`-]+)", full_text, re.I)
    if m and not first:
        first = m.group(2).title()
    if first and last:
        return first, last

    # 4) Uppercase line near the top, but not a city/state line
    lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]
    state_token = re.compile(r"\b[A-Z]{2}\b")  # e.g., DE, MN, CA
    for ln in lines[:12]:
        if any(ch.isdigit() for ch in ln):
            continue
        if "," in ln:
            continue
        if state_token.search(ln):
            continue
        toks = ln.split()
        if 2 <= len(toks) <= 3 and all(re.fullmatch(r"[A-Za-z'`-]+", t) for t in toks):
            upper_ratio = sum(t.isupper() for t in toks) / len(toks)
            if upper_ratio >= 2/3:
                return (toks[1].title(), toks[0].title())

    return first, last

# Address patterns
STREET_SUFFIX = r"(Street|St|Road|Rd|Avenue|Ave|Boulevard|Blvd|Lane|Ln|Court|Ct|Circle|Cir|Drive|Dr|Way|Wy|Terrace|Ter|Place|Pl)"
ADDR_LINE     = rf"(?:\b[0-9]\s+)?\b\d{{1,6}}\s+[A-Za-z0-9.'\- ]+?\s{STREET_SUFFIX}\b"
CITY_ST_ZIP   = r"([A-Za-z][A-Za-z .'-]+?),\s*([A-Z]{2})\s*(\d{5}(?:-\d{4})?)\b"

def parse_driver_license_text(full_text: str) -> dict:
    data = {
        "Phone": "",
        "Date": datetime.now().strftime("%m/%d/%Y"),
        "Have you Been Here Before?": "",
        "First Name": "",
        "Last Name": "",
        "Date Of Birth": "",
        "Email Address": "",
        "Address": "",
        "City": "",
        "State": "",
        "Zip Code": "",
        "Type Of Service": "",
        "Payment Type": "",
        "Amount": "",
        "Comments": "",
    }

    # Normalize whitespace
    text = re.sub(r"[ \t]+", " ", full_text)
    text = re.sub(r"\s*\n\s*", "\n", text)

    # Names
    fn_val, ln_val = _extract_names(text)
    if fn_val: data["First Name"] = fn_val
    if ln_val: data["Last Name"]  = ln_val

    # DOB
    m = re.search(r"(DOB|Date of Birth)[:\s]*([0-1]?\d/[0-3]?\d/\d{4})", text, re.IGNORECASE)
    if m: data["Date Of Birth"] = m.group(2)

    # Address line
    addr_line = re.search(ADDR_LINE, text, re.IGNORECASE)
    if addr_line:
        data["Address"] = addr_line.group(0).strip().title()

    # City, State, Zip
    csz = re.search(CITY_ST_ZIP, text)
    if csz:
        data["City"]     = csz.group(1).title()
        data["State"]    = csz.group(2)
        data["Zip Code"] = csz.group(3)

    # Phone (if present on some forms)
    ph = re.search(r"\b(\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4})\b", text)
    if ph: data["Phone"] = ph.group(1)

    if DEBUG:
        print(f"Parsed -> First: {data['First Name']} | Last: {data['Last Name']} | DOB: {data['Date Of Birth']}", flush=True)
        print(f"Parsed -> Address: {data['Address']}", flush=True)
        print(f"Parsed -> City/State/Zip: {data['City']} / {data['State']} / {data['Zip Code']}", flush=True)

    return data
# ==================================================


# ======================= Main =====================
def process_driver_license(image_path):
    full = detect_text(image_path)
    data = parse_driver_license_text(full)
    append_row(data, image_path)

if __name__ == "__main__":
    directory  = BASE_DIR
    image_path = sys.argv[1] if len(sys.argv) > 1 else find_latest_image(directory)
    process_driver_license(image_path)
# ==================================================
